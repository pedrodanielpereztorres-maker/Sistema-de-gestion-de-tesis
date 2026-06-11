import logging
import reflex as rx
from typing import List, Dict, Any
import csv
import io
from datetime import datetime
from ..database_manager import obtener_conexion
from fpdf import FPDF
from .estado_boveda import EstadoBoveda

logger = logging.getLogger(__name__)

class EstadoReportes(rx.State):
    resumen_global: Dict[str, int] = {
        "total_estudiantes": 0,
        "con_pasantia": 0,
        "sin_pasantia": 0,
        "total_tesis": 0
    }
    estadisticas_carreras: List[Dict[str, Any]] = []
    mejores_tutores: List[Dict[str, Any]] = []
    mejores_empresas: List[Dict[str, Any]] = []
    procesando: bool = False

    async def exportar_tesis_excel(self):
        """
        Genera y descarga un Excel con todas las tesis registradas de manera profesional.
        """
        boveda = await self.get_state(EstadoBoveda)
        if not boveda.lista_tesis:
            await boveda.cargar_tesis()
        
        if not boveda.lista_tesis:
            return rx.toast.warning("No hay tesis registradas para exportar.")
        
        try:
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Bóveda de Tesis"
            
            # Estilos profesionales
            color_primario = "6366F1" # Indigo
            color_secundario = "F1F5F9"
            
            fuente_titulo = Font(name='Arial', size=16, bold=True, color="FFFFFF")
            fuente_header = Font(name='Arial', size=12, bold=True, color="FFFFFF")
            fuente_normal = Font(name='Arial', size=11)
            
            alineacion_centro = Alignment(horizontal='center', vertical='center')
            alineacion_izq = Alignment(horizontal='left', vertical='center')
            
            relleno_titulo = PatternFill(start_color=color_primario, end_color=color_primario, fill_type="solid")
            relleno_header = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
            relleno_fila_par = PatternFill(start_color=color_secundario, end_color=color_secundario, fill_type="solid")
            
            borde_delgado = Border(
                left=Side(style='thin', color="CBD5E1"),
                right=Side(style='thin', color="CBD5E1"),
                top=Side(style='thin', color="CBD5E1"),
                bottom=Side(style='thin', color="CBD5E1")
            )
            
            # Encabezado del reporte
            ws.merge_cells('A1:J1')
            ws['A1'] = "SISTEMA DE GESTIÓN DE TESIS - REPORTE DE BÓVEDA DE TESIS"
            ws['A1'].font = fuente_titulo
            ws['A1'].alignment = alineacion_centro
            ws['A1'].fill = relleno_titulo
            
            ws.merge_cells('A2:J2')
            ws['A2'] = f"Descripción: Listado detallado de todos los trabajos de grado y tesis registrados. | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws['A2'].font = Font(name='Arial', size=11, italic=True)
            ws['A2'].alignment = alineacion_centro
            
            # Cabeceras de tabla
            headers = ["ID", "Cédula", "Nombre", "Apellido", "Carrera", "Tutor Académico", "Tutor Empresarial", "Empresa", "Título", "Pública"]
            for col_num, header in enumerate(headers, 1):
                celda = ws.cell(row=4, column=col_num, value=header)
                celda.font = fuente_header
                celda.alignment = alineacion_centro
                celda.fill = relleno_header
                celda.border = borde_delgado
            
            # Datos
            for row_num, t in enumerate(boveda.lista_tesis, 5):
                datos_fila = [
                    t.get("id", ""), t.get("cedula_estudiante", ""), t.get("nombre_estudiante", ""),
                    t.get("apellido_estudiante", ""), t.get("carrera", ""), t.get("tutor_academico", ""),
                    t.get("tutor_empresa", ""), t.get("nombre_empresa", ""),
                    t.get("titulo", ""), "Sí" if t.get("publico", False) else "No",
                ]
                for col_num, valor in enumerate(datos_fila, 1):
                    celda = ws.cell(row=row_num, column=col_num, value=valor)
                    celda.font = fuente_normal
                    celda.alignment = alineacion_izq if col_num > 2 else alineacion_centro
                    celda.border = borde_delgado
                    if row_num % 2 == 0:
                        celda.fill = relleno_fila_par
            
            # Ajustar anchos de columna
            anchos = [5, 12, 15, 15, 20, 20, 20, 20, 40, 10]
            for i, ancho in enumerate(anchos, 1):
                ws.column_dimensions[ws.cell(row=4, column=i).column_letter].width = ancho

            salida = io.BytesIO()
            wb.save(salida)
            return rx.download(
                data=salida.getvalue(),
                filename=f"reporte_boveda_{datetime.now().strftime('%d_%m_%Y')}.xlsx"
            )
        except Exception as e:
            logger.exception("Error al generar reporte de bóveda: %s", e)
            return rx.toast.error(f"Error al generar reporte: {e}")

    async def cargar_reportes(self):
        """Carga todos los datos analíticos desde la base de datos."""
        conn = obtener_conexion()
        if conn is None:
            return rx.toast.error("Error de conexión al cargar reportes.")
        
        try:
            with conn:
                with conn.cursor() as cursor:
                    # 1. Resumen Global
                    cursor.execute("SELECT COUNT(*) FROM estudiante WHERE esta_activo = TRUE")
                    total = cursor.fetchone()[0]
                    
                    cursor.execute("SELECT COUNT(*) FROM estudiante WHERE tutor_academico_id IS NOT NULL AND esta_activo = TRUE")
                    con_pasantia = cursor.fetchone()[0]
                    
                    cursor.execute("SELECT COUNT(*) FROM tesis")
                    total_tesis = cursor.fetchone()[0]
                    
                    self.resumen_global = {
                        "total_estudiantes": total,
                        "con_pasantia": con_pasantia,
                        "sin_pasantia": total - con_pasantia,
                        "total_tesis": total_tesis
                    }

                    # 2. Estadísticas por Carrera
                    cursor.execute("""
                        SELECT c.nombre, COUNT(e.id) 
                        FROM carrera c 
                        LEFT JOIN estudiante e ON c.id = e.carrera_id AND e.esta_activo = TRUE
                        WHERE c.esta_activa = TRUE
                        GROUP BY c.nombre
                    """)
                    resultados_carreras = cursor.fetchall()
                    max_est = max([r[1] for r in resultados_carreras]) if resultados_carreras else 1
                    self.estadisticas_carreras = [
                        {"nombre": r[0], "cantidad": r[1], "progreso": (r[1] / max_est) * 100}
                        for r in resultados_carreras
                    ]

                    # 3. Mejores Tutores (Top 5)
                    cursor.execute("""
                        SELECT ta.nombre || ' ' || ta.apellido, COUNT(e.id) as cantidad
                        FROM tutor_academico ta
                        JOIN estudiante e ON ta.id = e.tutor_academico_id
                        WHERE ta.esta_activo = TRUE
                        GROUP BY ta.id, ta.nombre, ta.apellido
                        ORDER BY cantidad DESC LIMIT 5
                    """)
                    self.mejores_tutores = [{"nombre": r[0], "cantidad": r[1]} for r in cursor.fetchall()]

                    # 4. Todas las Empresas (con conteo de pasantes)
                    # Correo y teléfono se toman del tutor empresarial asociado
                    cursor.execute("""
                        SELECT emp.nombre, emp.direccion, MAX(te.correo), MAX(te.telefono), COUNT(e.id) as cantidad
                        FROM empresa emp
                        LEFT JOIN tutor_empresarial te ON emp.id = te.empresa_id
                        LEFT JOIN estudiante e ON te.id = e.tutor_empresarial_id AND e.esta_activo = TRUE
                        GROUP BY emp.id, emp.nombre, emp.direccion
                        ORDER BY cantidad DESC, emp.nombre ASC
                    """)
                    self.mejores_empresas = [
                        {"nombre": r[0], "direccion": r[1], "correo": r[2], "telefono": r[3], "cantidad": r[4]} 
                        for r in cursor.fetchall()
                    ]
        except Exception as e:
            logger.exception("Error al cargar reportes: %s", e)
        finally:
            if conn:
                try:
                    conn.close()
                except Exception:
                    pass

    def exportar_empresas_excel(self):
        """Genera un reporte Excel profesional de todas las empresas vinculadas."""
        if not self.mejores_empresas:
            return rx.toast.warning("No hay datos de empresas para exportar.")
            
        try:
            import io
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Vinculación Empresarial"
            
            # Estilos
            color_primario = "10B981" # Emerald
            color_secundario = "F1F5F9"
            
            fuente_titulo = Font(name='Arial', size=16, bold=True, color="FFFFFF")
            fuente_header = Font(name='Arial', size=12, bold=True, color="FFFFFF")
            fuente_normal = Font(name='Arial', size=11)
            
            alineacion_centro = Alignment(horizontal='center', vertical='center')
            alineacion_izq = Alignment(horizontal='left', vertical='center')
            
            relleno_titulo = PatternFill(start_color=color_primario, end_color=color_primario, fill_type="solid")
            relleno_header = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
            relleno_fila_par = PatternFill(start_color=color_secundario, end_color=color_secundario, fill_type="solid")
            
            borde_delgado = Border(
                left=Side(style='thin', color="CBD5E1"),
                right=Side(style='thin', color="CBD5E1"),
                top=Side(style='thin', color="CBD5E1"),
                bottom=Side(style='thin', color="CBD5E1")
            )
            
            # Encabezado
            ws.merge_cells('A1:E1')
            ws['A1'] = "SISTEMA DE GESTIÓN DE TESIS - REPORTE DE VINCULACIÓN EMPRESARIAL"
            ws['A1'].font = fuente_titulo
            ws['A1'].alignment = alineacion_centro
            ws['A1'].fill = relleno_titulo
            
            ws.merge_cells('A2:E2')
            ws['A2'] = f"Descripción: Listado detallado de todas las entidades aliadas y su carga de pasantes. | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws['A2'].font = Font(name='Arial', size=11, italic=True)
            ws['A2'].alignment = alineacion_centro
            
            headers = ["Empresa", "Dirección", "Correo de Contacto", "Teléfono", "Pasantes Actuales"]
            for col_num, header in enumerate(headers, 1):
                celda = ws.cell(row=4, column=col_num, value=header)
                celda.font = fuente_header
                celda.alignment = alineacion_centro
                celda.fill = relleno_header
                celda.border = borde_delgado
            
            for row_num, emp in enumerate(self.mejores_empresas, 5):
                datos_fila = [emp["nombre"], emp["direccion"], emp["correo"], emp["telefono"], emp["cantidad"]]
                for col_num, valor in enumerate(datos_fila, 1):
                    celda = ws.cell(row=row_num, column=col_num, value=valor)
                    celda.font = fuente_normal
                    celda.alignment = alineacion_centro if col_num == 5 else alineacion_izq
                    celda.border = borde_delgado
                    if row_num % 2 == 0:
                        celda.fill = relleno_fila_par
            
            anchos = [35, 40, 30, 20, 15]
            for i, ancho in enumerate(anchos, 1):
                ws.column_dimensions[ws.cell(row=4, column=i).column_letter].width = ancho

            salida = io.BytesIO()
            wb.save(salida)
            return rx.download(
                data=salida.getvalue(),
                filename=f"reporte_empresas_{datetime.now().strftime('%d_%m_%Y')}.xlsx"
            )
        except Exception as e:
            logger.exception("Error al generar reporte excel: %s", e)
            return rx.toast.error("Error al generar el reporte de empresas en Excel.")

    def exportar_empresas_pdf(self):
        """Genera un reporte PDF profesional de las empresas."""
        if not self.mejores_empresas:
            return rx.toast.warning("No hay datos para generar el PDF.")

        try:
            pdf = FPDF(orientation='P', unit='mm', format='A4')
            pdf.add_page()
            
            # --- Encabezado con Identidad Institucional ---
            pdf.set_font("Helvetica", 'B', 18)
            pdf.set_text_color(31, 41, 55)  # Gris oscuro
            pdf.cell(0, 15, "SISTEMA DE GESTIÓN DE TESIS", ln=True, align='C')
            
            # Línea decorativa
            pdf.set_draw_color(99, 102, 241)  # Indigo
            pdf.set_line_width(0.5)
            pdf.line(10, 25, 200, 25)
            pdf.ln(5)

            # --- Título del Reporte y Metadatos ---
            pdf.set_font("Helvetica", 'B', 14)
            pdf.cell(0, 10, "Reporte de Vinculación Empresarial", ln=True, align='L')
            
            pdf.set_font("Helvetica", 'I', 9)
            pdf.set_text_color(100, 116, 139)  # Slate
            pdf.cell(0, 5, f"Fecha de emisión: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ln=True, align='L')
            pdf.ln(10)
            
            # --- Descripción del Reporte ---
            pdf.set_font("Helvetica", '', 10)
            pdf.set_text_color(30, 41, 59)
            desc = "Listado oficial de entidades receptoras de pasantes y vinculación académica institucional."
            pdf.cell(0, 10, desc, ln=True)
            
            # Resumen de métricas en el PDF
            pdf.set_font("Helvetica", 'B', 10)
            pdf.cell(0, 8, f"Total de Empresas Registradas: {len(self.mejores_empresas)}", ln=True)
            pdf.ln(10)

            # --- Cabecera de tabla ---
            pdf.set_fill_color(99, 102, 241) # Color Indigo del sistema
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", 'B', 10)
            
            pdf.cell(65, 10, "Nombre de la Empresa", 1, 0, 'C', True)
            pdf.cell(75, 10, "Correo de Contacto", 1, 0, 'C', True)
            pdf.cell(30, 10, "Teléfono", 1, 0, 'C', True)
            pdf.cell(20, 10, "Pasantes", 1, 1, 'C', True)
            
            # --- Datos de la tabla ---
            pdf.set_text_color(30, 41, 59)
            pdf.set_font("Helvetica", '', 9)
            
            relleno = False
            for emp in self.mejores_empresas:
                pdf.set_fill_color(248, 250, 252) if relleno else pdf.set_fill_color(255, 255, 255)
                
                # Control de longitud para evitar desbordamiento
                nombre = (emp["nombre"][:35] + '..') if len(emp["nombre"]) > 37 else emp["nombre"]
                correo_val = emp["correo"] or "N/A"
                correo_val = (correo_val[:40] + '..') if len(correo_val) > 42 else correo_val
                
                pdf.cell(65, 8, nombre, 1, 0, 'L', relleno)
                pdf.cell(75, 8, correo_val, 1, 0, 'L', relleno)
                pdf.cell(30, 8, emp["telefono"] or "N/A", 1, 0, 'C', relleno)
                pdf.cell(20, 8, str(emp["cantidad"]), 1, 1, 'C', relleno)
                relleno = not relleno

            # --- Pie de página ---
            pdf.set_y(-20)
            pdf.set_font("Helvetica", 'I', 8)
            pdf.set_text_color(148, 163, 184)
            pdf.cell(0, 10, f"Página {pdf.page_no()} - Documento Administrativo Confidencial", align='C')
            
            # Obtener el contenido de forma segura
            pdf_output = bytes(pdf.output())
            
            return rx.download(data=pdf_output, filename=f"reporte_empresas_{datetime.now().strftime('%d_%m_%Y')}.pdf")
        except Exception as e:
            logger.exception("Error al generar PDF de reportes: %s", e)
            return rx.toast.error("Error técnico al generar el PDF.")